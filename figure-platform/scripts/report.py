#!/usr/bin/env python3
"""
Generate a self-contained HTML progress report for the Interactive 3D Figures project.
Reads:
  - backend/human_eval/Result Tracking.xlsx  (human evaluation study)
  - backend/results/*.json                    (AI critic evaluations)
Outputs:
  - report.html  (self-contained, no external dependencies)

Usage:
  python3 scripts/report.py
  python3 scripts/report.py --output /path/to/report.html
"""

import json, glob, re, os, sys, argparse
from collections import defaultdict, Counter
from datetime import datetime

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE, 'backend', 'human_eval', 'Result Tracking.xlsx')
RESULTS_DIR = os.path.join(BASE, 'backend', 'results')
FIGURES_DIR = os.path.join(BASE, '..', 'figures')

METRICS = ['Geometry Accuracy', 'Interactivity & Usability', 'Faithfulness', 'Label Quality', 'Concept Accuracy']
METRIC_SHORT = ['Geom', 'Interact', 'Faith', 'Labels', 'Concept']

# ── Helpers ───────────────────────────────────────────────────────────────────
def safe_avg(lst):
    vals = [v for v in lst if isinstance(v, (int, float))]
    return round(sum(vals) / len(vals), 2) if vals else None

def score_color(s):
    if s is None: return '#aaa'
    if s >= 4.0: return '#1a5c2a'
    if s >= 3.0: return '#7a4a00'
    return '#8b1a1a'

def score_class(s):
    if s is None: return 'score score-na'
    if s >= 4.0: return 'score score-hi'
    if s >= 3.0: return 'score score-mid'
    return 'score score-lo'

def score_bg(s):
    return ''

def bar(val, max_val=5, color='#1a1a1a', height=6):
    pct = min(100, (val or 0) / max_val * 100)
    return f'<span class="bar-track"><span class="bar-fill" style="width:{pct:.0f}%"></span></span>'

def extract_iferror_val(cell_val):
    if isinstance(cell_val, (int, float)): return cell_val
    if isinstance(cell_val, str):
        m = re.search(r'COMPUTED_VALUE"""\),["\s]*(.*?)["\s]*\)', cell_val)
        if m:
            v = m.group(1).strip().strip('"')
            try: return float(v)
            except: return v if v else None
    return cell_val

# ── Load human eval data ──────────────────────────────────────────────────────
def load_human_eval():
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Data']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data = []
    for r in rows:
        if not r[1] or not isinstance(r[1], str): continue
        week = str(r[0])[:10] if r[0] else None
        scores = {METRICS[i]: r[4+i] for i in range(5)}
        scores['Visual Aesthetics'] = r[9]
        overall = safe_avg([r[4+i] for i in range(5)])
        failures = [f.strip() for f in r[10].split(',')] if r[10] and isinstance(r[10], str) else []
        data.append({
            'week': week, 'prompt': r[1], 'model': r[2], 'figure': r[3],
            'scores': scores, 'overall': overall,
            'failures': failures, 'notes': r[11],
            'rendering_failure': overall is None,
        })

    # Model comparison tab
    ws_mc = wb['Model Comparison']
    model_table = []
    prompt_filter = ws_mc.cell(2,1).value or 'all'
    for row in ws_mc.iter_rows(min_row=2, max_row=40, values_only=True):
        model = extract_iferror_val(row[1])
        if not model or not isinstance(model, str): continue
        overall = extract_iferror_val(row[2])
        if not isinstance(overall, float): continue
        metric_scores = [extract_iferror_val(row[i]) for i in range(3, 8)]
        top_fails = extract_iferror_val(row[9])
        model_table.append({'model': model, 'overall': overall,
                             'metrics': metric_scores, 'top_failures': top_fails,
                             'prompt_filter': prompt_filter})

    # Prompt comparison tab
    ws_pc = wb['Prompt Comparison']
    prompt_table = []
    model_filter = ws_pc.cell(2,1).value or 'all'
    for row in ws_pc.iter_rows(min_row=2, max_row=20, values_only=True):
        prompt = extract_iferror_val(row[1])
        if not prompt or not isinstance(prompt, str): continue
        overall = extract_iferror_val(row[2])
        if not isinstance(overall, float): continue
        metric_scores = [extract_iferror_val(row[i]) for i in range(3, 8)]
        top_fails = extract_iferror_val(row[9])
        prompt_table.append({'prompt': prompt, 'overall': overall,
                              'metrics': metric_scores, 'top_failures': top_fails,
                              'model_filter': model_filter})

    return data, model_table, prompt_table

# ── Load AI critic results ────────────────────────────────────────────────────
def load_ai_results():
    results = []
    for f in sorted(glob.glob(os.path.join(RESULTS_DIR, '*.json'))):
        try:
            d = json.load(open(f))
            ev = d.get('evaluation') or {}
            if not ev: continue
            stem = os.path.splitext(d.get('filename',''))[0]
            results.append({
                'filename': d.get('filename'),
                'stem': stem,
                'model': d.get('model') or 'gpt-4o',
                'experiment': d.get('experiment') or 'baseline',
                'overall': ev.get('overall_average'),
                'geometry': ev.get('geometry_accuracy'),
                'interactivity': ev.get('interactivity_usability'),
                'faithfulness': ev.get('faithfulness'),
                'label_quality': ev.get('label_quality'),
                'concept_accuracy': ev.get('concept_accuracy'),
                'failures': ev.get('failure_modes', []),
                'notes': ev.get('notes',''),
            })
        except: pass
    return results

# ── Count corpus ──────────────────────────────────────────────────────────────
def count_corpus():
    if not os.path.isdir(FIGURES_DIR): return 0, 0
    chapters = [d for d in os.listdir(FIGURES_DIR) if os.path.isdir(os.path.join(FIGURES_DIR, d))]
    images = []
    for ch in chapters:
        for ext in ['*.png','*.jpg','*.jpeg']:
            images += glob.glob(os.path.join(FIGURES_DIR, ch, ext))
    return len(chapters), len(images)

# ── HTML generation ───────────────────────────────────────────────────────────
def html_header():
    return '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Interactive 3D Figures — Progress Report</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=EB+Garamond:ital,wght@0,400;0,600;0,700;1,400&family=Source+Code+Pro:wght@400;600&display=swap');
  *{box-sizing:border-box;margin:0;padding:0}
  body{font-family:Georgia,"Times New Roman",serif;background:#fff;color:#1a1a1a;font-size:11pt;line-height:1.65}
  .page{max-width:780px;margin:0 auto;padding:72px 64px}

  /* ── Cover ── */
  .cover{border-bottom:2px solid #1a1a1a;padding-bottom:36px;margin-bottom:48px}
  .cover-title{font-size:24pt;font-weight:700;line-height:1.2;letter-spacing:-.02em;margin-bottom:10px}
  .cover-subtitle{font-size:11pt;color:#555;font-style:italic;margin-bottom:24px}
  .cover-meta{display:flex;gap:48px;font-size:9.5pt;color:#444;border-top:1px solid #ccc;padding-top:14px;margin-top:20px}
  .cover-meta span{display:flex;flex-direction:column;gap:2px}
  .cover-meta strong{font-size:8pt;text-transform:uppercase;letter-spacing:.1em;color:#888;font-weight:600}

  /* ── Stats strip ── */
  .stats-strip{display:flex;gap:0;border:1px solid #ccc;margin-bottom:40px}
  .stat-cell{flex:1;padding:14px 18px;border-right:1px solid #ccc;text-align:center}
  .stat-cell:last-child{border-right:none}
  .stat-num{font-size:22pt;font-weight:700;line-height:1;display:block}
  .stat-label{font-size:8pt;text-transform:uppercase;letter-spacing:.08em;color:#666;display:block;margin-top:3px}
  .stat-sub{font-size:8pt;color:#999;display:block;margin-top:1px;font-style:italic}

  /* ── Headings ── */
  h2{font-size:13pt;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:#1a1a1a;
     margin:52px 0 16px;padding-bottom:4px;border-bottom:1.5px solid #1a1a1a}
  h3{font-size:11pt;font-weight:700;color:#1a1a1a;margin:20px 0 8px}

  /* ── Body text ── */
  p{margin-bottom:10px}
  ol,ul{margin-left:20px}
  li{margin-bottom:3px}

  /* ── Tables ── */
  table{width:100%;border-collapse:collapse;margin-bottom:20px;font-size:9.5pt}
  thead tr{border-top:1.5px solid #1a1a1a;border-bottom:1px solid #1a1a1a}
  th{font-size:8pt;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#1a1a1a;
     padding:7px 10px;text-align:left;background:none}
  td{padding:6px 10px;border-bottom:1px solid #ddd;vertical-align:middle;font-size:9.5pt}
  tbody tr:last-child td{border-bottom:1.5px solid #1a1a1a}
  tfoot td{font-weight:600;border-top:1px solid #aaa;font-size:9pt}

  /* ── Score display ── */
  .score{font-variant-numeric:tabular-nums;font-weight:600}
  .score-hi{color:#1a5c2a}
  .score-mid{color:#7a4a00}
  .score-lo{color:#8b1a1a}
  .score-na{color:#aaa}

  /* ── Callout boxes ── */
  .callout{border-left:3px solid #1a1a1a;padding:8px 14px;margin:14px 0;font-size:9.5pt}
  .callout p{margin:0}
  .callout.note{border-color:#2c5282;background:#f7f9fc}
  .callout.warn{border-color:#7a4a00;background:#fdf6ec}
  .callout.good{border-color:#1a5c2a;background:#f4faf5}

  /* ── Tags / code ── */
  .model-name{font-family:"Source Code Pro","Courier New",monospace;font-size:8.5pt;color:#1a1a1a}
  .prompt-name{font-family:"Source Code Pro","Courier New",monospace;font-size:8.5pt;color:#2c5282}
  .tag{font-family:"Source Code Pro","Courier New",monospace;font-size:8pt;color:#444;margin-right:4px;white-space:nowrap}
  code{font-family:"Source Code Pro","Courier New",monospace;font-size:8.5pt;color:#1a1a1a}

  /* ── Layout ── */
  .grid2{display:grid;grid-template-columns:1fr 1fr;gap:28px;margin-bottom:20px}
  .grid3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:20px;margin-bottom:20px}
  .section{border:1px solid #ccc;padding:18px 22px;margin-bottom:20px}

  /* ── Bar charts ── */
  .bar-track{height:6px;background:#e8e8e8;width:100%;display:inline-block;vertical-align:middle}
  .bar-fill{height:6px;background:#1a1a1a;display:block}

  /* ── Footer ── */
  .doc-footer{margin-top:64px;padding-top:12px;border-top:1px solid #ccc;
    font-size:8.5pt;color:#888;display:flex;justify-content:space-between}

  /* ── Print ── */
  @media print{
    body{font-size:10pt}
    .page{padding:48px 52px;max-width:100%}
    h2{page-break-before:auto}
    table{page-break-inside:avoid}
    .section{page-break-inside:avoid}
    .cover{page-break-after:always}
  }
</style>
</head>
<body><div class="page">
'''

def html_footer():
    return f'''
<div class="doc-footer">
  <span>Interactive 3D Figures Project &mdash; Internal Progress Report</span>
  <span>Generated {datetime.now().strftime("%B %d, %Y at %H:%M")}</span>
</div>
</div></body></html>'''

def render_score_badge(s):
    if s is None: return '<span class="score score-na">—</span>'
    return f'<span class="{score_class(s)}">{s:.2f}</span>'

def render_metric_cells(metrics, padding='6px 10px'):
  cells = ''
  for m in metrics:
    if isinstance(m, (int, float)):
      cells += f'<td style="text-align:right;padding:{padding}">{render_score_badge(float(m))}</td>'
    else:
      cells += f'<td style="text-align:right;padding:{padding}" class="score score-na">—</td>'
  return cells

def render_failure_tags(top_fails):
    if not top_fails or not isinstance(top_fails, str): return ''
    tags = ', '.join(f'<span class="tag">{f.strip()}</span>' for f in top_fails.split(','))
    return tags

def section_title():
    return f'''
<div class="cover">
  <div class="cover-title">Interactive 3D Figures<br>Progress Report</div>
  <div class="cover-subtitle">Agentic LLM pipeline for generating interactive 3D visualizations from 2D textbook figures</div>
  <div class="cover-meta">
    <span><strong>Date</strong>{datetime.now().strftime("%B %d, %Y")}</span>
    <span><strong>Study period</strong>February 9 – March 9, 2026</span>
    <span><strong>Status</strong>Active — data collection phase</span>
    <span><strong>Document type</strong>Internal progress report</span>
  </div>
</div>
'''

def section_background():
    return '''
<h2>1. Background and Motivation</h2>

<h3>Goal</h3>
<p>Educational textbooks in computer vision and related fields rely heavily on static 2D diagrams to explain spatial,
geometric, and optical concepts. These diagrams are effective on the page but offer no means for a reader to explore
perspective, rotate a scene, or interact with the underlying geometry. The goal of this project is to automatically
convert these static 2D figures into <strong>self-contained interactive 3D visualizations</strong> that run in the
browser, allowing students to manipulate camera angles, toggle labels, adjust parameters, and develop genuine
spatial intuition for the concepts being taught.</p>

<p>The target corpus is a full computer vision textbook with 2,377 source figures across 54 chapters. Rather than
hand-authoring visualizations for each figure, the project pursues a <strong>fully automated conversion pipeline</strong>
driven by large language models, with the aim of producing high-quality interactive content at scale.</p>

<h3>Core problem</h3>
<p>Direct single-shot LLM generation of interactive 3D scenes proves unreliable in practice. Three failure classes
dominate:</p>
<ol style="margin:8px 0 12px 20px;line-height:1.9">
  <li><strong>Spatial reasoning errors.</strong> Models consistently mis-specify 3D coordinates, vertex ordering, surface
  normals, and camera parameters. The mapping from a 2D diagram to a geometrically correct 3D scene requires precise
  numerical reasoning that current models handle inconsistently.</li>
  <li><strong>Concept hallucination.</strong> Models sometimes generate plausible-looking but factually incorrect
  representations of the underlying concept — introducing geometry, labels, or interactive elements that do not
  correspond to the source figure or its pedagogical purpose.</li>
  <li><strong>Rendering fragility.</strong> Generated Three.js code frequently fails to render due to API misuse,
  missing scene setup, or logic errors that are not caught until execution. Approximately 15% of single-shot
  outputs produce blank or broken scenes.</li>
</ol>

<h3>Research approach</h3>
<p>Rather than attempting to solve these problems through a single improved prompt, the project takes an
<strong>empirical failure-mode analysis approach</strong>:</p>
<ul style="margin:8px 0 12px 20px;line-height:1.9">
  <li>Systematically collect evaluation data across diverse models, prompt variants, and figures, recording which
  failure modes occur and under what conditions.</li>
  <li>Operate under the working assumption that <strong>failure modes are finite and enumerable</strong>. If the
  complete failure taxonomy can be characterised, targeted interventions (prompt engineering, retrieval,
  verification passes, fine-tuning) can be designed for each category.</li>
  <li>Use the failure-mode frequency data to prioritise which problems to solve first, focusing effort on the
  highest-impact categories rather than addressing edge cases.</li>
</ul>

<h2>2. Pipeline Architecture and Current Stage</h2>

<h3>Current pipeline (implemented)</h3>
<p>The system currently implements a three-stage agentic loop:</p>

<div style="margin:16px 0 20px;padding:16px 20px;border:1px solid #ccc">
  <div style="display:flex;align-items:flex-start;gap:0;font-size:9.5pt">
    <div style="flex:1;padding-right:16px;border-right:1px solid #ddd">
      <div style="font-weight:700;text-transform:uppercase;font-size:8pt;letter-spacing:.07em;color:#555;margin-bottom:6px">1 &mdash; Generator</div>
      <p style="margin:0">An LLM receives the source figure image plus a structured prompt (optionally including chapter QMD source and a Three.js base scaffold). Outputs a complete, self-contained HTML file with embedded Three.js scene.</p>
    </div>
    <div style="flex:1;padding:0 16px;border-right:1px solid #ddd">
      <div style="font-weight:700;text-transform:uppercase;font-size:8pt;letter-spacing:.07em;color:#555;margin-bottom:6px">2 &mdash; Critic</div>
      <p style="margin:0">A separate LLM instance evaluates the output against five quality metrics and a structured failure-mode taxonomy (10+ categories). Returns a JSON critique with scores and specific defect descriptions.</p>
    </div>
    <div style="flex:1;padding-left:16px">
      <div style="font-weight:700;text-transform:uppercase;font-size:8pt;letter-spacing:.07em;color:#555;margin-bottom:6px">3 &mdash; Refinement loop</div>
      <p style="margin:0">The critic feedback is injected into the next generation prompt. The loop repeats until a quality threshold is reached or a maximum round count is exceeded. Human evaluators periodically score final outputs.</p>
    </div>
  </div>
</div>

<h3>Current stage: transitioning from manual to automated experiments</h3>
<p>The evaluation study documented in this report was conducted <strong>primarily through manual experimentation</strong>:
models and prompts were selected and run by hand, outputs were scored by a single human evaluator per session, and
results were logged in a tracking spreadsheet. This approach yielded the 153-row dataset analysed below.</p>
<p>The project is now transitioning toward <strong>fully automated batch evaluation</strong> via the <span class="model-name">agent.js</span>
CLI, which orchestrates the Generator–Critic loop without human intervention and logs structured JSON results
automatically. This shift is intended to enable evaluation at the scale of the full 2,377-figure corpus.</p>

<div class="callout note">
  <p><strong>Why automation matters.</strong> Running a single model–prompt combination across the 9-figure
  evaluation set currently takes approximately 20–30 minutes of manual effort. Scaling to 100 figures across
  12 model variants would require hundreds of hours of manual work. The automated loop reduces this to a
  background batch job.</p>
</div>

<h3>Future pipeline extensions (planned)</h3>
<p>Two extensions are under consideration to address the core failure modes identified in this study:</p>
<ul style="margin:8px 0 12px 20px;line-height:1.9">
  <li><strong>Interaction Planner agent.</strong> A dedicated upstream stage that parses the chapter source (QMD/LaTeX)
  and the figure&#39;s pedagogical context to determine <em>what</em> should be interactive, <em>what parameters</em> the
  user should be able to control, and which labels and annotations are factually required. This agent would produce a
  structured specification that the Generator follows, reducing concept hallucination and label errors. A verification
  pass would check for factual inconsistencies before passing the spec to the Generator.</li>
  <li><strong>Interaction quality oracle.</strong> A specialised critic module focused solely on interaction design:
  whether controls are within bounds, whether the interaction reveals the intended concept, and whether the initial
  camera view is pedagogically correct. This addresses the two highest-frequency failure modes (Initial-View-Wrong
  and Missing/Bad Labels).</li>
</ul>
'''

def section_overview(data, ai_results, n_chapters, n_images):
    n_human = len(data)
    n_rendering_fail = sum(1 for d in data if d['rendering_failure'])
    n_scored = n_human - n_rendering_fail
    n_models = len(set(d['model'] for d in data))
    n_prompts = len(set(d['prompt'] for d in data))
    n_figures_eval = len(set(d['figure'] for d in data))
    n_ai = len(ai_results)
    weeks = sorted(set(d['week'] for d in data if d['week']))

    return f'''
<h2>3. Study at a Glance</h2>

<div class="stats-strip">
  <div class="stat-cell">
    <span class="stat-num">{n_images:,}</span>
    <span class="stat-label">Source figures</span>
    <span class="stat-sub">{n_chapters} chapters</span>
  </div>
  <div class="stat-cell">
    <span class="stat-num">{n_human}</span>
    <span class="stat-label">Human evaluations</span>
    <span class="stat-sub">{n_scored} scored · {n_rendering_fail} render fail</span>
  </div>
  <div class="stat-cell">
    <span class="stat-num">{n_models}</span>
    <span class="stat-label">Models tested</span>
    <span class="stat-sub">{n_prompts} prompt variants</span>
  </div>
  <div class="stat-cell">
    <span class="stat-num">{n_figures_eval}</span>
    <span class="stat-label">Figures evaluated</span>
    <span class="stat-sub">{len(weeks)} weeks of study</span>
  </div>
  <div class="stat-cell">
    <span class="stat-num">{n_ai}</span>
    <span class="stat-label">AI critic scores</span>
    <span class="stat-sub">automated pipeline</span>
  </div>
</div>

<p>The evaluation study ran from February 9 to March 9, 2026 across {n_models} models and {n_prompts} prompt variants.
All {n_human} human evaluations were conducted on {n_figures_eval} figures drawn from two chapters of the corpus
(homographies and imaging). The {n_ai} AI critic scores provide an automated baseline against the same figure set.
Corpus coverage currently stands at {n_figures_eval/n_images*100:.1f}% of the {n_images:,}-figure corpus.</p>
'''

def section_human_model(model_table, data):
    if not model_table:
        return '<h2>2. Human Evaluation — Model Comparison</h2><p>No model comparison data available.</p>'

    sorted_models = sorted(model_table, key=lambda x: x['overall'], reverse=True)
    prompt_filter = sorted_models[0].get('prompt_filter','') if sorted_models else ''

    # rendering failure rates from raw data
    fail_rates = {}
    model_counts = defaultdict(lambda: {'total': 0, 'failed': 0})
    for d in data:
        model_counts[d['model']]['total'] += 1
        if d['rendering_failure']:
            model_counts[d['model']]['failed'] += 1
    for m, c in model_counts.items():
        fail_rates[m] = c['failed'] / c['total'] * 100 if c['total'] else 0

    rows = ''
    for i, m in enumerate(sorted_models):
        rank = i + 1
        fr = fail_rates.get(m['model'], 0)
        fr_str = f'<span class="score score-lo">{fr:.0f}%</span>' if fr > 10 else f'{fr:.0f}%'
        rows += f'''<tr>
          <td style="font-weight:700;color:#999;font-size:9pt">{rank}</td>
          <td><span class="model-name">{m["model"]}</span></td>
          <td style="text-align:right">{render_score_badge(m["overall"])}</td>
          {render_metric_cells(m["metrics"])}
          <td style="text-align:right">{fr_str}</td>
          <td style="font-size:8.5pt">{render_failure_tags(m["top_failures"])}</td>
        </tr>'''

    return f'''
<h2>4. Human Evaluation — Model Comparison</h2>
<p style="margin-bottom:12px">Scores are human ratings on a 1–5 scale. Prompt filter: <span class="prompt-name">{prompt_filter}</span>. Models ranked by overall average across all five quality dimensions.</p>
<table>
  <thead><tr>
    <th>#</th><th>Model</th><th style="text-align:right">Overall</th>
    <th style="text-align:right">Geom.</th>
    <th style="text-align:right">Interact.</th>
    <th style="text-align:right">Faith.</th>
    <th style="text-align:right">Labels</th>
    <th style="text-align:right">Concept</th>
    <th style="text-align:right">Rend. fail</th>
    <th>Primary failures</th>
  </tr></thead>
  <tbody>{rows}</tbody>
</table>
<div class="callout good"><p><strong>Best model.</strong> <span class="model-name">{sorted_models[0]["model"]}</span> achieves {sorted_models[0]["overall"]:.2f}/5 overall. Models with the <code>-with-base-code</code> suffix receive a Three.js scaffold as additional input context, consistently improving geometric accuracy and rendering reliability.</p></div>
<div class="callout warn"><p><strong>Baseline.</strong> <span class="model-name">gpt-4o</span> scores 1.16/5 — the gap between the baseline and best model ({sorted_models[0]["overall"]-1.16:.2f} points) illustrates the gains from newer model families and prompt engineering.</p></div>
'''

def section_human_prompt(prompt_table):
    if not prompt_table:
        return '<h2>3. Human Evaluation — Prompt Comparison</h2><p>No prompt data.</p>'

    sorted_prompts = sorted(prompt_table, key=lambda x: x['overall'], reverse=True)
    model_filter = sorted_prompts[0].get('model_filter','') if sorted_prompts else ''

    rows = ''
    for i, p in enumerate(sorted_prompts):
        rows += f'''<tr>
          <td style="font-weight:700;color:#999;font-size:9pt">{i+1}</td>
          <td><span class="prompt-name">{p["prompt"]}</span></td>
          <td style="text-align:right">{render_score_badge(p["overall"])}</td>
          {render_metric_cells(p["metrics"])}
          <td style="font-size:8.5pt">{render_failure_tags(p["top_failures"])}</td>
        </tr>'''

    prompt_descriptions = {
        'with_qmd': 'Full prompt + QMD source chapter context',
        'with_reqs': 'Full prompt + explicit requirements list',
        'one_line': 'Single-line minimal prompt',
        'limit_qmd': 'QMD context with length-limited prompt',
    }

    desc_rows = ''.join(
        f'<tr><td><span class="prompt-name">{k}</span></td><td style="font-size:12px;color:#555">{v}</td></tr>'
        for k, v in prompt_descriptions.items()
    )

    return f'''
<h2>5. Human Evaluation — Prompt Comparison</h2>
<p style="margin-bottom:12px">Scores averaged across all models and figures. Model filter: all models.</p>
<table>
  <thead><tr>
    <th>#</th><th>Prompt variant</th><th style="text-align:right">Overall</th>
    <th style="text-align:right">Geom.</th>
    <th style="text-align:right">Interact.</th>
    <th style="text-align:right">Faith.</th>
    <th style="text-align:right">Labels</th>
    <th style="text-align:right">Concept</th>
    <th>Primary failures</th>
  </tr></thead>
  <tbody>{rows}</tbody>
</table>

<h3>Prompt descriptions</h3>
<table style="margin-bottom:20px">
  <thead><tr><th>Variant</th><th>Description</th></tr></thead>
  <tbody>{desc_rows}</tbody>
</table>

<div class="callout good"><p><strong>with_qmd achieves the highest concept accuracy (4.44)</strong> — supplying the source chapter text as context helps the model understand the pedagogical intent of each figure, leading to more accurate conceptual representations.</p></div>
<div class="callout warn"><p><strong>limit_qmd underperforms despite including chapter context</strong> — truncating the QMD source to fit a length limit removes critical information. Providing the full chapter text is preferable to a truncated version.</p></div>
'''

def section_weekly_trend(data):
    week_scores = defaultdict(lambda: {'scores': [], 'models': set(), 'failed': 0, 'total': 0})
    for d in data:
        if not d['week']: continue
        week_scores[d['week']]['total'] += 1
        week_scores[d['week']]['models'].add(d['model'])
        if d['rendering_failure']:
            week_scores[d['week']]['failed'] += 1
        elif d['overall'] is not None:
            week_scores[d['week']]['scores'].append(d['overall'])

    rows = ''
    max_avg = max((safe_avg(v['scores']) or 0) for v in week_scores.values())
    for week in sorted(week_scores):
        v = week_scores[week]
        avg = safe_avg(v['scores'])
        n_scored = len(v['scores'])
        models_str = ', '.join(f'<span class="model-name">{m}</span>' for m in sorted(v['models']))
        bar_html = bar(avg or 0, max_val=5, height=6) if avg else ''
        fail_str = f'<span class="score score-lo">{v["failed"]}</span>' if v['failed'] > 0 else '0'
        rows += f'''<tr>
          <td>{week}</td>
          <td style="text-align:right">{v["total"]}</td>
          <td style="text-align:right">{n_scored}</td>
          <td style="text-align:right">{fail_str}</td>
          <td style="text-align:right">{render_score_badge(avg)}&nbsp;&nbsp;{bar_html}</td>
          <td style="font-size:8.5pt">{models_str}</td>
        </tr>'''

    return f'''
<h2>6. Weekly Evaluation Progress</h2>
<table>
  <thead><tr>
    <th>Week of</th><th style="text-align:right">Evaluations</th>
    <th style="text-align:right">Scored</th>
    <th style="text-align:right">Render fails</th>
    <th style="text-align:right">Avg. score</th>
    <th>Models evaluated</th>
  </tr></thead>
  <tbody>{rows}</tbody>
</table>
<div class="callout note"><p>The week of Mar 2 achieves the highest average (3.94), reflecting strong performance of <span class="model-name">claude-opus-4.6</span> and <span class="model-name">gemini-3.1-pro</span> paired with the <span class="prompt-name">with_qmd</span> prompt. The Mar 9 dip is attributable to inclusion of the <span class="model-name">gpt-4o</span> baseline, which scores approximately 1.0.</p></div>
'''

def section_failure_modes(data):
    fail_counter = Counter()
    per_model_fails = defaultdict(Counter)
    for d in data:
        for f in d['failures']:
            if f:
                fail_counter[f] += 1
                per_model_fails[d['model']][f] += 1

    total_evals = len(data)
    rows = ''
    for mode, count in fail_counter.most_common(15):
        pct = count / total_evals * 100
        bar_html = bar(pct, max_val=100, height=6)
        rows += f'''<tr>
          <td><span class="tag">{mode}</span></td>
          <td style="text-align:right;font-weight:700">{count}</td>
          <td style="text-align:right">{pct:.0f}%</td>
          <td style="width:160px">{bar_html}</td>
        </tr>'''

    # categorize
    spatial = ['Bad-3D','Initial-View-Wrong','Bad-Labels','Missing-Labels']
    interaction = ['Interaction-OOB','Interaction-Broken','Interaction-Missing','Interaction-Unintuitive']
    faithfulness = ['Original-Mismatch','Hallucination','Concept-Misunderstanding','Rendering-Failure']

    cat_rows = ''
    for cat_name, cat_modes in [('Spatial / Geometry', spatial), ('Interaction', interaction), ('Faithfulness & Correctness', faithfulness)]:
        cat_total = sum(fail_counter[m] for m in cat_modes)
        items = ''.join(f'<tr><td><span class="tag">{m}</span></td><td style="text-align:right">{fail_counter[m]}</td></tr>' for m in cat_modes)
        cat_rows += f'<tr><td colspan="2" style="padding-top:12px;font-weight:700;font-size:9pt;text-transform:uppercase;letter-spacing:.05em;border-bottom:none;color:#555">{cat_name}&nbsp;&nbsp;<span style="font-weight:400">({cat_total})</span></td></tr>{items}'

    return f'''
<h2>6. Failure Mode Analysis</h2>
<div class="grid2">
  <div>
    <h3>By category</h3>
    <table>
      <thead><tr><th>Failure mode</th><th style="text-align:right">Count</th></tr></thead>
      <tbody>{cat_rows}</tbody>
    </table>
  </div>
  <div>
    <h3>Full ranking (n={total_evals} evaluations)</h3>
    <table>
      <thead><tr><th>Failure mode</th><th style="text-align:right">Count</th><th style="text-align:right">%</th><th>Frequency</th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
</div>
<div class="callout note"><p><strong>Label quality is the most persistent problem.</strong> "Missing-Labels" and "Bad-Labels" combined account for 91 occurrences (59% of evaluations). "Initial-View-Wrong" (51 occurrences, 33%) suggests models default to suboptimal initial camera angles despite explicit prompt instructions.</p></div>
'''

def section_per_figure(data):
    fig_scores = defaultdict(lambda: {'scores': [], 'models': set(), 'prompts': set()})
    for d in data:
        if d['overall'] is not None:
            fig_scores[d['figure']]['scores'].append(d['overall'])
            fig_scores[d['figure']]['models'].add(d['model'])
            fig_scores[d['figure']]['prompts'].add(d['prompt'])

    rows = ''
    for fig in sorted(fig_scores, key=lambda f: -(safe_avg(fig_scores[f]['scores']) or 0)):
        v = fig_scores[fig]
        avg = safe_avg(v['scores'])
        chapter, name = fig.split('/') if '/' in fig else ('', fig)
        rows += f'''<tr>
          <td style="color:#888">{chapter}</td>
          <td style="font-weight:600"><span class="model-name">{name}</span></td>
          <td style="text-align:right">{render_score_badge(avg)}</td>
          <td style="text-align:right">{len(v["scores"])}</td>
          <td style="text-align:right">{len(v["models"])}</td>
        </tr>'''

    return f'''
<h2>7. Per-Figure Analysis</h2>
<p style="margin-bottom:12px">Nine figures evaluated across all model and prompt combinations. Averages exclude rendering failures.</p>
<table>
  <thead><tr>
    <th>Chapter</th><th>Figure</th><th style="text-align:right">Avg. score</th>
    <th style="text-align:right">Scored runs</th>
    <th style="text-align:right">Models</th>
  </tr></thead>
  <tbody>{rows}</tbody>
</table>
'''

def section_ai_critic(ai_results):
    if not ai_results:
      return '<h2>8. AI Critic Evaluation</h2><p>No AI critic results found in results/ directory.</p>'

    rows = ''
    for r in sorted(ai_results, key=lambda x: -(x['overall'] or 0)):
        metrics = [r.get('geometry'), r.get('interactivity'), r.get('faithfulness'), r.get('label_quality'), r.get('concept_accuracy')]
        fail_tags = ' '.join(f'<span class="tag">{f}</span>' for f in (r.get('failures') or []))
        rows += f'''<tr style="vertical-align:top">
          <td style="font-size:9pt;font-weight:600;padding:4px 8px;max-width:140px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="{r['filename']}">{r["filename"]}</td>
          <td style="padding:4px 8px"><span class="model-name">{r["model"]}</span></td>
          <td style="text-align:center;padding:4px 8px">{render_score_badge(r["overall"])}</td>
          {render_metric_cells(metrics, padding='4px 8px')}
          <td style="font-size:8pt;max-width:260px;white-space:normal;line-height:1.6;padding:4px 8px">{fail_tags}</td>
        </tr>'''

    ai_avg = safe_avg([r['overall'] for r in ai_results if r['overall']])
    return f'''
<h2>8. AI Critic Evaluation (Automated Pipeline)</h2>
<p style="margin-bottom:12px">Results scored by the automated critic module via <code>critic.js</code> — no human involvement. Group average: {ai_avg:.2f}/5.</p>
<table style="font-size:9pt;table-layout:fixed;width:100%">
  <thead><tr>
    <th style="padding:5px 8px">Figure</th><th style="padding:5px 8px">Generator model</th><th style="text-align:right;padding:5px 8px">Overall</th>
    <th style="text-align:right;padding:5px 8px">Geom.</th>
    <th style="text-align:right;padding:5px 8px">Interact.</th>
    <th style="text-align:right;padding:5px 8px">Faith.</th>
    <th style="text-align:right;padding:5px 8px">Labels</th>
    <th style="text-align:right;padding:5px 8px">Concept</th>
    <th style="padding:5px 8px">Failure modes</th>
  </tr></thead>
  <tbody>{rows}</tbody>
</table>
<div class="callout warn"><p>AI critic scores cluster at 1.8–2.2/5 for the gpt-4o baseline — broadly consistent with human evaluation of the same model (1.16/5 human average). The automated critic scores slightly higher, suggesting it may underpenalise subtle geometric and faithfulness errors that human raters detect.</p></div>
'''

def section_comparison(data, ai_results):
    # Find overlapping figures: human has imaging/brdf etc, AI has brdf.png
    ai_by_stem = {r['stem']: r for r in ai_results}

    overlaps = []
    for d in data:
        if d['overall'] is None: continue
        fig_stem = d['figure'].split('/')[-1] if '/' in d['figure'] else d['figure']
        if fig_stem in ai_by_stem:
            overlaps.append((d, ai_by_stem[fig_stem]))

    if not overlaps:
        return ''

    rows = ''
    diffs = []
    for human_d, ai_r in overlaps:
        h_score = human_d['overall']
        a_score = ai_r['overall']
        diff = round(a_score - h_score, 2) if (h_score and a_score) else None
        diffs.append(diff)
        diff_str = (f'<span class="score score-lo">+{diff:.2f}</span>' if diff and diff > 0
                    else f'<span class="score score-hi">{diff:.2f}</span>' if diff else '—')
        rows += f'''<tr>
          <td><span class="model-name">{human_d["figure"].split("/")[-1]}</span></td>
          <td><span class="model-name">{human_d["model"]}</span></td>
          <td><span class="prompt-name">{human_d["prompt"]}</span></td>
          <td style="text-align:right">{render_score_badge(h_score)}</td>
          <td style="text-align:right">{render_score_badge(a_score)}</td>
          <td style="text-align:right">{diff_str}</td>
        </tr>'''

    mean_diff = safe_avg([d for d in diffs if d is not None])
    bias_note = ''
    if mean_diff is not None:
        if mean_diff > 0.3:
            bias_note = f'<div class="callout warn"><p>AI critic scores average +{mean_diff:.2f} higher than human scores on the same figures, indicating the automated critic is systematically lenient. The AI critic does not penalise subtle conceptual errors the way human evaluators do.</p></div>'
        elif mean_diff < -0.3:
            bias_note = f'<div class="callout good"><p>AI critic scores average {mean_diff:.2f} lower than human scores — the automated critic is conservative relative to human raters.</p></div>'
        else:
            bias_note = f'<div class="callout good"><p>AI and human scores are closely aligned (mean difference: {mean_diff:+.2f}), suggesting the automated critic pipeline is well-calibrated for this evaluation regime.</p></div>'

    return f'''
<h2>9. Human vs. AI Critic Comparison</h2>
<p style="margin-bottom:12px">{len(set(d["figure"].split("/")[-1] for d, _ in overlaps))} figures evaluated by both human raters and the automated critic. AI critic records are gpt-4o baseline runs with no prompt metadata.</p>
<table>
  <thead><tr>
    <th>Figure</th><th>Model</th><th>Prompt</th>
    <th style="text-align:right">Human score</th>
    <th style="text-align:right">AI critic</th>
    <th style="text-align:right">Difference</th>
  </tr></thead>
  <tbody>{rows}</tbody>
</table>
{bias_note}
'''

def section_findings():
    return '''
<h2>10. Key Findings and Recommendations</h2>

<div class="grid2">
  <div>
    <h3>Strengths</h3>
    <ul style="margin-left:18px;line-height:1.9;font-size:10pt">
      <li><strong>Geometry generation</strong> is the strongest metric across all models (avg 4.2+).</li>
      <li>The <span class="prompt-name">with_qmd</span> prompt significantly improves concept accuracy — source chapter context is highly beneficial.</li>
      <li><span class="model-name">claude-opus-4.6-with-base-code</span> achieves 4.28/5 overall with a Three.js scaffold.</li>
      <li><span class="model-name">gpt-5.3-codex</span> is the best balanced model (3.96/5) with strong interactivity scores.</li>
      <li>Newer model generations (5.x series) dramatically outperform the gpt-4o baseline (1.16 → 3.9+).</li>
      <li>The automated critic pipeline is directionally consistent with human evaluation.</li>
    </ul>
  </div>
  <div>
    <h3>Persistent challenges</h3>
    <ul style="margin-left:18px;line-height:1.9;font-size:10pt">
      <li><strong>Label quality</strong> is the weakest metric — models frequently omit or mislabel annotations.</li>
      <li><strong>Initial camera view</strong> is frequently suboptimal despite explicit prompt instructions.</li>
      <li><strong>Rendering failures</strong> remain a reliability concern (~15% of runs).</li>
      <li>Interactivity is often out-of-bounds or unintuitive despite explicit requirements.</li>
      <li>Faithfulness to the original figure is inconsistent (Original-Mismatch: 45 occurrences).</li>
    </ul>
  </div>
</div>

<h3>Recommended next steps</h3>
<ol style="margin-left:20px;line-height:1.9;font-size:10pt">
  <li><strong>Scale evaluation.</strong> Run the agent loop on 50+ figures across 5+ chapters; measure score distribution at corpus scale.</li>
  <li><strong>Agent loop ablation.</strong> Compare round-1 vs. round-2 vs. round-3 outputs to quantify the gains from iterative refinement.</li>
  <li><strong>Label quality intervention.</strong> Apply targeted prompt engineering focused on annotation accuracy, then re-evaluate.</li>
  <li><strong>Inter-annotator agreement.</strong> Have two or more evaluators score the same outputs to validate rubric reliability and compute Cohen\xe2\x80\x99s \xce\xba.</li>
  <li><strong>AI critic calibration.</strong> Expand the human/AI overlap set to compute Pearson correlation per metric and establish correction factors.</li>
  <li><strong>Chapter coverage map.</strong> Track which of the 54 chapters have been converted and at what quality level, targeting systematic coverage.</li>
  <li><strong>Rendering reliability.</strong> Instrument the pipeline to automatically log, classify, and retry rendering failures.</li>
</ol>
'''

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--output', default=os.path.join(BASE, 'report.html'))
    args = parser.parse_args()

    print("Loading human evaluation data...")
    data, model_table, prompt_table = load_human_eval()
    print(f"  {len(data)} rows loaded")

    print("Loading AI critic results...")
    ai_results = load_ai_results()
    print(f"  {len(ai_results)} AI-scored results")

    print("Counting corpus...")
    n_chapters, n_images = count_corpus()
    print(f"  {n_images} images across {n_chapters} chapters")

    print("Generating report...")
    html = (
        html_header() +
        section_title() +
        section_background() +
        section_overview(data, ai_results, n_chapters, n_images) +
        section_human_model(model_table, data) +
        section_human_prompt(prompt_table) +
        section_failure_modes(data) +
        section_per_figure(data) +
        section_ai_critic(ai_results) +
        section_comparison(data, ai_results) +
        section_findings() +
        html_footer()
    )

    with open(args.output, 'w') as f:
        f.write(html)
    print(f"\n✓ Report saved → {args.output}")

if __name__ == '__main__':
    main()
